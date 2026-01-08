from fastapi import APIRouter, HTTPException, Body, Depends
from sqlalchemy.orm import Session
from typing import Dict, Any, List
from ..db.database import get_db
from ..models.company import Company
from ..models.education_facility import EducationFacility
from ..models.certification_center import CertificationCenter
from ..models.location import Location
from ..models.language import Language
from ..models.category import Category
from ..models.skill import Skill
from ..schemas.common import SuccessResponse
import json
import os
from sqlalchemy import and_
from openpyxl import load_workbook

router = APIRouter(prefix="/data-management", tags=["Data Management"])

COMPANIES_FILE = "app/utils/files/it_companies.json"
CERTIFICATIONS_FILE = "app/utils/files/it_certifications.json"
EDUCATION_FILE = "app/utils/files/education_institutions_corrected.json"
COUNTRIES_FILE = "app/utils/files/countries.json"
LANGUAGES_FILE = "app/utils/files/languages.json"
CATEGORIES_FILE = "app/utils/files/Categories.xlsx"

def read_json_file(file_path: str) -> List[Dict[str, Any]]:
    """Read JSON file and return its content with error handling"""
    try:
        with open(file_path, 'r', encoding='utf-8') as f:
            return json.load(f)
    except FileNotFoundError:
        return []
    except json.JSONDecodeError as e:
        return []
    except Exception as e:
        return []

def batch_insert_companies(db: Session, companies_data: List[Dict[str, Any]]) -> int:
    """Batch insert companies for better performance"""
    companies_imported = 0
    
    unique_companies = {}
    for company_data in companies_data:
        key = (company_data.get("name"), company_data.get("country"), company_data.get("icon"))
        if key not in unique_companies:
            unique_companies[key] = company_data
    
    existing_companies = db.query(Company).filter(
        and_(
            Company.is_deleted == False
        )
    ).all()
    
    existing_keys = {(c.name, c.country, c.icon) for c in existing_companies}
    
    new_companies = []
    for company_data in unique_companies.values():
        company_key = (company_data.get("name"), company_data.get("country"), company_data.get("icon"))
        if company_key not in existing_keys:
            company = Company(
                name=company_data.get("name"),
                icon=company_data.get("icon"),
                country=company_data.get("country")
            )
            new_companies.append(company)
            companies_imported += 1
    
    if new_companies:
        db.bulk_save_objects(new_companies)
        db.commit()
    
    return companies_imported

def batch_insert_education_facilities(db: Session, education_data: List[Dict[str, Any]]) -> int:
    """Batch insert education facilities for better performance"""
    education_imported = 0
    
    unique_facilities = {}
    for edu_data in education_data:
        key = (edu_data.get("name"), edu_data.get("country"), edu_data.get("icon"))
        if key not in unique_facilities:
            unique_facilities[key] = edu_data
    
    existing_facilities = db.query(EducationFacility).filter(
        and_(
            EducationFacility.is_deleted == False
        )
    ).all()
    
    existing_keys = {(f.name, f.country, f.icon) for f in existing_facilities}
    
    new_facilities = []
    for edu_data in unique_facilities.values():
        facility_key = (edu_data.get("name"), edu_data.get("country"), edu_data.get("icon"))
        if facility_key not in existing_keys:
            education_facility = EducationFacility(
                name=edu_data.get("name"),
                icon=edu_data.get("icon"),
                country=edu_data.get("country")
            )
            new_facilities.append(education_facility)
            education_imported += 1
    
    if new_facilities:
        db.bulk_save_objects(new_facilities)
        db.commit()
    
    return education_imported

def batch_insert_certification_centers(db: Session, certifications_data: List[Dict[str, Any]]) -> int:
    """Batch insert certification centers for better performance"""
    certifications_imported = 0
    
    unique_certifications = {}
    for cert_data in certifications_data:
        key = (cert_data.get("name"), cert_data.get("icon"))
        if key not in unique_certifications:
            unique_certifications[key] = cert_data
    
    existing_certifications = db.query(CertificationCenter).filter(
        and_(
            CertificationCenter.is_deleted == False
        )
    ).all()
    
    existing_keys = {(c.name, c.icon) for c in existing_certifications}
    
    new_certifications = []
    for cert_data in unique_certifications.values():
        cert_key = (cert_data.get("name"), cert_data.get("icon"))
        if cert_key not in existing_keys:
            certification_center = CertificationCenter(
                name=cert_data.get("name"),
                icon=cert_data.get("icon")
            )
            new_certifications.append(certification_center)
            certifications_imported += 1
    
    if new_certifications:
        db.bulk_save_objects(new_certifications)
        db.commit()
    
    return certifications_imported

def batch_insert_locations(db: Session, countries_data: List[Dict[str, Any]]) -> int:
    """Batch insert locations for better performance"""
    locations_imported = 0
    
    unique_locations = {}
    for country_data in countries_data:
        key = (country_data.get("name"), country_data.get("code"))
        if key not in unique_locations:
            unique_locations[key] = country_data
    
    existing_locations = db.query(Location).filter(
        and_(
            Location.is_deleted == False
        )
    ).all()
    
    existing_keys = {(l.name, l.code) for l in existing_locations}
    
    new_locations = []
    for country_data in unique_locations.values():
        location_key = (country_data.get("name"), country_data.get("code"))
        if location_key not in existing_keys:
            location = Location(
                name=country_data.get("name"),
                code=country_data.get("code"),
                phone_code=country_data.get("phone_code"),
                flag_image=country_data.get("flag_image")
            )
            new_locations.append(location)
            locations_imported += 1
    
    if new_locations:
        db.bulk_save_objects(new_locations)
        db.commit()
    
    return locations_imported

def batch_insert_languages(db: Session, languages_data: List[Dict[str, Any]]) -> int:
    """Batch insert languages for better performance"""
    languages_imported = 0
    
    existing_languages = db.query(Language).all()
    existing_names = {lang.name for lang in existing_languages}
    
    new_languages = []
    for lang_data in languages_data:
        language_name = lang_data.get("language")
        if language_name and language_name not in existing_names:
            language = Language(name=language_name)
            new_languages.append(language)
            languages_imported += 1
    
    if new_languages:
        db.bulk_save_objects(new_languages)
        db.commit()
    
    return languages_imported

def batch_insert_categories_and_skills(db: Session) -> Dict[str, int]:
    """
    Batch insert categories, sub-categories, and skills from Excel file
    Returns dictionary with counts of imported items
    """
    categories_imported = 0
    subcategories_imported = 0
    skills_imported = 0
    
    try:
        workbook = load_workbook(CATEGORIES_FILE, read_only=True)
        sheet = workbook.active
        
        existing_categories = db.query(Category).filter(Category.is_active == True).all()
        existing_categories_dict = {cat.name: cat for cat in existing_categories}
        
        existing_skills = db.query(Skill).filter(Skill.is_deleted == False).all()
        existing_skills_set = {skill.name.strip().lower() for skill in existing_skills}
        
        new_categories = []
        new_subcategories = []
        new_skills = []
        new_skills_set = set()
        
        current_category_name = None
        
        for row in sheet.iter_rows(min_row=3, values_only=True):
            if not row or len(row) < 4:
                continue
                
            category_name = row[1]
            subcategory_name = row[2]
            skills_text = row[3]
            
            if category_name:
                current_category_name = str(category_name).strip()
            
            if not current_category_name or not subcategory_name or not skills_text:
                continue
            
            category_name = current_category_name
            
            category_name = str(category_name).strip()
            if category_name not in existing_categories_dict:
                if not any(cat.name == category_name for cat in new_categories):
                    parent_category = Category(
                        name=category_name,
                        is_active=True,
                        parent_id=None
                    )
                    new_categories.append(parent_category)
                    categories_imported += 1
            
            subcategory_name = str(subcategory_name).strip()
            subcategory_exists = any(
                cat.name == subcategory_name and 
                cat.parent and cat.parent.name == category_name 
                for cat in existing_categories
            )
            
            if not subcategory_exists:
                subcategory_exists = any(
                    cat.name == subcategory_name and 
                    hasattr(cat, '_parent_name') and 
                    cat._parent_name == category_name 
                    for cat in new_subcategories
                )
            
            if not subcategory_exists:
                subcategory = Category(
                    name=subcategory_name,
                    is_active=True,
                    parent_id=None
                )
                subcategory._parent_name = category_name
                new_subcategories.append(subcategory)
                subcategories_imported += 1
            
            skills_list = [s.strip() for s in str(skills_text).split(',')]
            for skill_name in skills_list:
                skill_name = skill_name.strip()
                if not skill_name:
                    continue
                
                skill_name_lower = skill_name.lower()
                
                if skill_name_lower not in existing_skills_set and skill_name_lower not in new_skills_set:
                    skill = Skill(
                        name=skill_name,
                        is_deleted=False
                    )
                    new_skills.append(skill)
                    new_skills_set.add(skill_name_lower)
                    skills_imported += 1
        
        if new_categories:
            db.add_all(new_categories)
            db.commit()
            
            for cat in new_categories:
                db.refresh(cat)
                existing_categories_dict[cat.name] = cat
        
        if new_subcategories:
            for subcat in new_subcategories:
                parent = existing_categories_dict.get(subcat._parent_name)
                if parent:
                    subcat.parent_id = parent.id
            
            db.add_all(new_subcategories)
            db.commit()
        
        if new_skills:
            db.bulk_save_objects(new_skills)
            db.commit()
        
        workbook.close()
        
        return {
            "categories_imported": categories_imported,
            "subcategories_imported": subcategories_imported,
            "skills_imported": skills_imported,
            "total_imported": categories_imported + subcategories_imported + skills_imported
        }
    
    except Exception as e:
        db.rollback()
        raise Exception(f"Failed to import from Excel: {str(e)}")

@router.post("/import-categories-skills", response_model=SuccessResponse)
async def import_categories_and_skills(db: Session = Depends(get_db)):
    """
    Import categories, sub-categories, and skills from Excel file (Categories.xlsx)
    """
    try:
        result = batch_insert_categories_and_skills(db)
        
        return SuccessResponse(
            msg="Categories and skills imported successfully",
            data=result
        )
    
    except Exception as e:
        db.rollback()
        raise HTTPException(status_code=500, detail=f"Failed to import categories and skills: {str(e)}")

@router.post("/import-all", response_model=SuccessResponse)
async def import_all_data(db: Session = Depends(get_db)):
    """
    Import all data from JSON files into database models with optimized batch operations
    """
    try:
        companies_data = read_json_file(COMPANIES_FILE)
        certifications_data = read_json_file(CERTIFICATIONS_FILE)
        education_data = read_json_file(EDUCATION_FILE)
        countries_data = read_json_file(COUNTRIES_FILE)
        languages_data = read_json_file(LANGUAGES_FILE)
        
        companies_imported = batch_insert_companies(db, companies_data)
        
        education_imported = batch_insert_education_facilities(db, education_data)
        
        certifications_imported = batch_insert_certification_centers(db, certifications_data)
        
        locations_imported = batch_insert_locations(db, countries_data)
        
        languages_imported = batch_insert_languages(db, languages_data)
        
        categories_result = batch_insert_categories_and_skills(db)
        
        total_imported = (companies_imported + education_imported + certifications_imported + 
                         locations_imported + languages_imported + categories_result["total_imported"])
        
        return SuccessResponse(
            msg="Data imported successfully",
            data={
                "companies_imported": companies_imported,
                "education_facilities_imported": education_imported,
                "certification_centers_imported": certifications_imported,
                "locations_imported": locations_imported,
                "languages_imported": languages_imported,
                "categories_imported": categories_result["categories_imported"],
                "subcategories_imported": categories_result["subcategories_imported"],
                "skills_imported": categories_result["skills_imported"],
                "total_imported": total_imported
            }
        )
    
    except Exception as e:
        db.rollback()
        raise HTTPException(status_code=500, detail=f"Failed to import data: {str(e)}")

@router.get("/stats", response_model=SuccessResponse)
async def get_import_stats(db: Session = Depends(get_db)):
    """
    Get statistics about imported data with optimized queries
    """
    try:
        companies_count = db.query(Company).filter(Company.is_deleted == False).count()
        education_count = db.query(EducationFacility).filter(EducationFacility.is_deleted == False).count()
        certifications_count = db.query(CertificationCenter).filter(CertificationCenter.is_deleted == False).count()
        locations_count = db.query(Location).filter(Location.is_deleted == False).count()
        languages_count = db.query(Language).count()
        categories_count = db.query(Category).filter(Category.is_active == True).count()
        skills_count = db.query(Skill).filter(Skill.is_deleted == False).count()
        
        companies_data = read_json_file(COMPANIES_FILE)
        education_data = read_json_file(EDUCATION_FILE)
        certifications_data = read_json_file(CERTIFICATIONS_FILE)
        countries_data = read_json_file(COUNTRIES_FILE)
        languages_data = read_json_file(LANGUAGES_FILE)
        
        return SuccessResponse(
            msg="Import statistics retrieved successfully",
            data={
                "database_stats": {
                    "companies": companies_count,
                    "education_facilities": education_count,
                    "certification_centers": certifications_count,
                    "locations": locations_count,
                    "languages": languages_count,
                    "categories": categories_count,
                    "skills": skills_count,
                    "total": companies_count + education_count + certifications_count + locations_count + languages_count + categories_count + skills_count
                },
                "json_file_stats": {
                    "companies": len(companies_data),
                    "education_facilities": len(education_data),
                    "certification_centers": len(certifications_data),
                    "locations": len(countries_data),
                    "languages": len(languages_data),
                    "total": len(companies_data) + len(education_data) + len(certifications_data) + len(countries_data) + len(languages_data)
                }
            }
        )
    
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Failed to get stats: {str(e)}") 
